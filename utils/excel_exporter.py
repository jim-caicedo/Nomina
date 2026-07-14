"""
Exportador de nómina a Excel.
Genera archivos Excel con formato profesional para liquidaciones de nómina.
"""

from datetime import date
from typing import List
from models.domain.nomina import RegistroNomina


def exportar_nomina_a_excel(
    registros: List[RegistroNomina],
    periodo_inicio: date,
    periodo_cierre: date,
    ruta_archivo: str = None,
) -> str:
    """
    Exporta los registros de nómina a un archivo Excel.

    Args:
        registros: Lista de registros de nómina a exportar
        periodo_inicio: Fecha de inicio del período
        periodo_cierre: Fecha de cierre del período
        ruta_archivo: Ruta donde guardar el archivo (opcional)

    Returns:
        Ruta del archivo Excel generado

    Raises:
        ImportError: Si openpyxl no está instalado
        Exception: Si hay error al generar el archivo
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError(
            "openpyxl no está instalado. Ejecuta: pip install openpyxl"
        )

    # Crear libro de trabajo
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Eliminar hoja por defecto

    # Bordes delgados comunes para los datos
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Estilo común para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # ========== HOJA 1: RESUMEN ==========
    ws_resumen = wb.create_sheet("Resumen")

    # Título
    ws_resumen.merge_cells("A1:E1")
    ws_resumen["A1"] = f"RESUMEN DE NÓMINA"
    ws_resumen["A1"].font = Font(name="Arial", size=16, bold=True)
    ws_resumen["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Período
    ws_resumen.merge_cells("A2:E2")
    ws_resumen["A2"] = (
        f"Período: {periodo_inicio.strftime('%d/%m/%Y')} - "
        f"{periodo_cierre.strftime('%d/%m/%Y')}"
    )
    ws_resumen["A2"].font = Font(name="Arial", size=12, bold=True)
    ws_resumen["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Calcular totales
    total_empleados = len(registros)
    total_devengado = sum(r.total_devengado for r in registros)
    total_deducciones = sum(r.total_deducciones for r in registros)
    total_neto = sum(r.salario_neto for r in registros)

    # Datos de resumen
    datos_resumen = [
        ["", "", "", "", ""],
        ["Total Empleados:", total_empleados, "", "", ""],
        ["Total Devengado:", f"${total_devengado:,.2f}", "", "", ""],
        ["Total Deducciones:", f"${total_deducciones:,.2f}", "", "", ""],
        ["Nómina Neta:", f"${total_neto:,.2f}", "", "", ""],
    ]

    for row_idx, row_data in enumerate(datos_resumen, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_resumen.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 1:
                cell.font = Font(name="Arial", size=11, bold=True)
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.font = Font(name="Arial", size=11)
                cell.alignment = Alignment(horizontal="left")

    # Ajustar ancho de columnas
    ws_resumen.column_dimensions["A"].width = 20
    ws_resumen.column_dimensions["B"].width = 20

    # ========== HOJA 2: DETALLE POR EMPLEADO ==========
    ws_detalle = wb.create_sheet("Detalle")

    # Título
    ws_detalle.merge_cells("A1:L1")
    ws_detalle["A1"] = "DETALLE DE LIQUIDACIÓN POR EMPLEADO"
    ws_detalle["A1"].font = Font(name="Arial", size=14, bold=True)
    ws_detalle["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Período
    ws_detalle.merge_cells("A2:L2")
    ws_detalle["A2"] = (
        f"Período: {periodo_inicio.strftime('%d/%m/%Y')} - "
        f"{periodo_cierre.strftime('%d/%m/%Y')}"
    )
    ws_detalle["A2"].font = Font(name="Arial", size=11, bold=True)
    ws_detalle["A2"].alignment = Alignment(horizontal="center", vertical="center")

    # Encabezados de tabla
    headers_detalle = [
        "Nº",
        "ID Empleado",
        "Periodo Inicio",
        "Periodo Cierre",
        "Días Laborados",
        "Salario Base",
        "Auxilio Transporte",
        "Horas Extras",
        "Total Devengado",
        "Descuento AFP",
        "Descuento EPS",
        "Salario Neto",
    ]

    # Escribir encabezados
    for col_idx, header in enumerate(headers_detalle, start=1):
        cell = ws_detalle.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Escribir datos
    for row_idx, registro in enumerate(registros, start=5):
        datos_fila = [
            row_idx - 4,  # Nº
            registro.empleado_id,
            registro.periodo_inicio.strftime("%d/%m/%Y"),
            registro.periodo_cierre.strftime("%d/%m/%Y"),
            registro.dias_laborados,
            registro.salario_base_periodo,
            registro.auxilio_transporte_periodo,
            registro.horas_extras,
            registro.total_devengado,
            registro.descuento_afp,
            registro.descuento_eps,
            registro.salario_neto,
        ]

        for col_idx, valor in enumerate(datos_fila, start=1):
            cell = ws_detalle.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin_border
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar ancho de columnas
    column_widths = [6, 12, 15, 15, 12, 15, 18, 12, 15, 12, 12, 15]
    for col_idx, width in enumerate(column_widths, start=1):
        ws_detalle.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ========== HOJA 3: ARCHIVO BANCARIO ==========
    ws_banco = wb.create_sheet("Archivo Bancario")

    # Encabezados de tabla bancaria (17 columnas)
    headers_banco = [
        "Cuenta Destino",
        "Nit Beneficiario",
        "Nombre Beneficiario",
        "Cod. Transaccion",
        "Tipo de cargo",
        "Valor Neto Pago",
        "No. Factura",
        "No. Control de pago",
        "Valor Retención en la Fuente",
        "Valor IVA",
        "Fecha Pago",
        "Número Nota Débito",
        "Valor Nota Débito",
        "Cod. Banco",
        "Tipo Cuenta",
        "Tipo Documento",
        "Inf. Adicional",
    ]

    # Escribir encabezados en Hoja Bancaria
    for col, texto in enumerate(headers_banco, start=1):
        celda = ws_banco.cell(row=1, column=col)
        celda.value = texto
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = header_alignment

    # Escribir la información bancaria de los empleados
    for fila, registro in enumerate(registros, start=2):
        emp = getattr(registro, "empleado", None)
        
        # Mapear los campos de acuerdo al script de SQLite proporcionado
        numero_cuenta = getattr(emp, "numero_cuenta", "") if emp else ""
        cedula_emp = getattr(emp, "cedula", "") if emp else ""
        nombre_completo = f"{getattr(emp, 'nombre', '')} {getattr(emp, 'apellido', '')}".strip() if emp else ""
        cargo_emp = getattr(emp, "cargo", "") if emp else ""
        codigo_banco = getattr(emp, "codigo_banco", "") if emp else ""
        tipo_cuenta = getattr(emp, "tipo_cuenta", "") if emp else ""
        tipo_documento = getattr(emp, "tipo_documento", "") if emp else ""

        # Distribución limpia alineada 1 a 1 con las 17 columnas superiores
        datos_banco = [
            numero_cuenta,                                                      # 1. Cuenta Destino
            cedula_emp,                                                         # 2. Nit Beneficiario (Cédula)
            nombre_completo if nombre_completo else f"Empleado {registro.empleado_id}", # 3. Nombre Beneficiario
            "902",                                                                 # 4. Cod. Transaccion
            cargo_emp,                                                          # 5. Tipo de cargo (Cargo)
            registro.salario_neto,                                              # 6. Valor Neto Pago
            "",                                                                 # 7. No. Factura
            "",                                                                 # 8. No. Control de pago
            0,                                                                  # 9. Valor Retención en la Fuente
            0,                                                                  # 10. Valor IVA
            registro.periodo_cierre.strftime("%d/%m/%Y"),                       # 11. Fecha Pago
            "",                                                                 # 12. Número Nota Débito
            "",                                                                 # 13. Valor Nota Débito
            codigo_banco,                                                       # 14. Cod. Banco
            tipo_cuenta,                                                        # 15. Tipo Cuenta
            tipo_documento,                                                     # 16. Tipo Documento
            "",                                                                 # 17. Inf. Adicional
        ]

        for columna, valor in enumerate(datos_banco, start=1):
            celda = ws_banco.cell(row=fila, column=columna, value=valor)
            celda.border = thin_border
            celda.font = Font(name="Arial", size=10)
            
            # Formatear alineaciones según el tipo de dato
            if columna in [1, 2, 6, 9, 10, 13]:  # Cuentas, Cédulas, Valores numéricos
                celda.alignment = Alignment(horizontal="right", vertical="center")
            else:  # Textos y códigos cortos
                celda.alignment = Alignment(horizontal="left", vertical="center")

    # Ajustar ancho de columnas de la Hoja Bancaria
    anchos_banco = [
        20, 18, 35, 18, 18, 18, 18, 20, 18, 12, 16, 18, 18, 12, 15, 18, 30
    ]

    for columna, ancho in enumerate(anchos_banco, start=1):
        ws_banco.column_dimensions[
            openpyxl.utils.get_column_letter(columna)
        ].width = ancho

    # Generar nombre de archivo si no se proporciona
    if ruta_archivo is None:
        ruta_archivo = f"nomina_{periodo_inicio.strftime('%Y%m%d')}_"
        ruta_archivo += f"{periodo_cierre.strftime('%Y%m%d')}.xlsx"

    # Guardar archivo
    wb.save(ruta_archivo)
    return ruta_archivo


def exportar_empleados_a_excel(
    empleados: list,
    ruta_archivo: str = None,
) -> str:
    """
    Exporta la lista de empleados a un archivo Excel.

    Args:
        empleados: Lista de diccionarios con datos de empleados (to_dict())
        ruta_archivo: Ruta donde guardar el archivo (opcional)

    Returns:
        Ruta del archivo Excel generado
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise ImportError("openpyxl no está instalado.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empleados"

    # Título
    ws.merge_cells("A1:J1")
    ws["A1"] = "LISTADO DE EMPLEADOS"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Fecha de exportación: {date.today().strftime('%d/%m/%Y')}"
    ws["A2"].font = Font(name="Arial", size=11)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Encabezados
    headers = [
        "ID", "Nombre", "Apellido", "Cargo", "Salario",
        "EPS", "AFP", "Sede", "Correo", "Teléfono"
    ]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Datos
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row_idx, emp in enumerate(empleados, start=5):
        fila = [
            emp.get("id", ""),
            emp.get("nombre", ""),
            emp.get("apellido", ""),
            emp.get("cargo", ""),
            emp.get("salario", 0),
            emp.get("eps", ""),
            emp.get("afp", ""),
            emp.get("sede_laboral", ""),
            emp.get("correo", ""),
            emp.get("telefono", ""),
        ]
        for col_idx, valor in enumerate(fila, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = thin
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="left")

    # Ancho de columnas
    anchos = [6, 16, 16, 18, 14, 18, 18, 16, 24, 14]
    for col_idx, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = ancho

    if ruta_archivo is None:
        ruta_archivo = f"empleados_{date.today().strftime('%Y%m%d')}.xlsx"

    wb.save(ruta_archivo)
    return ruta_archivo